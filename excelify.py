import os, re, time, dateutil, openpyxl
from openpyxl import Workbook
from sys import argv

script, file = argv

def load_convo():
	source = open(file,'r').read()
	#message_pattern = re.compile(r"\d+\/\d+\/\d+, .*",re.MULTILINE)
	#TODO - Fix this regex so that it matches the full message including line breaks

	conversation = re.findall(r"\d+\/\d+\/\d+, .*",source)
	print len(conversation)

	print '%d messages loaded...' % len(conversation)
	print 'Writing Messages to Excel file...'
	wb = Workbook()
	newfile = file.split('.')[0] + '.xlsx'
	write_messages(conversation, wb)
	write_links(conversation,wb)
	wb.save(newfile)

def write_messages(conversation, wb):
	ws = wb.create_sheet(index=0,title='Messages')
	ws.append(['Type','Datetime','Sender','Message'])
	#TODO: Add ,'# Words','# Links','Media'

	#Make pretty
	ws.freeze_panes = 'A2'
	column_widths = [10,15,20,100]
	for i, column_width in enumerate(column_widths):
		ws.column_dimensions[chr(65+i)].width = column_width

	#Parse strings
	p_message = re.compile("\d+\/\d+\/\d+, \d+:\d+ .*?:")
	i=0
	for line in conversation:
		i=i+1
		print '%d) %s' % (i, line)

		if p_message.match(line) is None:
			line_type = 'Event'
		else:
			line_type = 'Message'

		try:
			line_date = line[0:line.find('-',0)]
		except:
			line_date = "ERROR"

		try:
			line_sender = unicode(line[line.find('- ',0)+2:line.find(':',18)], 'utf8').encode('ascii', 'ignore')
		except:
			line_sender = ""

		if line_type == "Message":
			message = unicode(line[line.find(':',20)+2:], 'utf8').encode('ascii', 'ignore')
		else:
			message = unicode(line[line.find('- ')+2:], 'utf8').encode('ascii', 'ignore')
			line_sender = "--WhatsApp--"


		ws.append([line_type,line_date,line_sender,message])

	print ('-'*40)
	print '%d Messages exported' % len(conversation)

def write_links(conversation,wb):
	ws = wb.create_sheet(index=1,title='Links')
	ws.append(['Date','Sender','Link'])
	#TODO Extract Links

load_convo()